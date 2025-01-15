import argparse
import copy

import openpyxl

def get_base_state(cell, inherit:list=["value"]) -> dict:
	return {name: copy.copy(getattr(cell, name)) for name in inherit}

def assign_state(cell, res:dict):
	for name, state in res.items():
		setattr(cell, name, state)

def auto_unmerge_sheet(sheet, inherit:list=["value"]):
	merged_cells = list(sheet.merged_cells)

	for merged_cell in merged_cells:
		left, top, right, bottom = merged_cell.bounds
		value = get_base_state(merged_cell.start_cell, inherit)
		sheet.unmerge_cells(merged_cell.coord)
		flag_base = True

		for row in range(top, bottom+1):
			for col in range(left, right+1):
				if flag_base:
					flag_base = False
					continue
				assign_state(sheet.cell(row=row, column=col), value)

def auto_unmerge_workbook(workbook, inherit:list=["value"]):
	for sheet in workbook:
		auto_unmerge_sheet(sheet, inherit)

if __name__ == '__main__':
	parser = argparse.ArgumentParser(description="Auto unmerge all the merged cells in xlsx file with specific style retain.")
	parser.add_argument("FILE", type=str, help="Excel file to unmerge.")
	parser.add_argument("-o", "--output", default="unmerged.xlsx", type=str, help="Where to save the unmerged result.")
	parser.add_argument("-s", "--inherit-state", action="extend", nargs="+", default=["value"], choices=["value", "alignment", "border", "data_type", "fill", "font"], help="Which style retain, default is value.")
	args = parser.parse_args()

	workbook = openpyxl.load_workbook(args.FILE)
	auto_unmerge_workbook(workbook, inherit=args.inherit_state)
	workbook.save(args.output)
